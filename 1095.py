from calendar import monthrange
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
import openpyxl
import re
from datetime import date

#Core class which populates and updates GUI  
class MainApplication(tk.Frame):

    #Initialize main tkinter window
    def __init__(self, window) -> None:
        self.company = None
        self.window = window
        self.window.title("1095-C")
        self.window.geometry("1000x650")
        self.window.resizable(False, False)

        #Window header
        window_label = Label(text = "HAR FINANCIAL")
        window_label.place(anchor = CENTER, relx = .5, rely = 0.041)
        window_label.config(font = ("Segoe UI bold", 15))
        window_label = Label(text = "Affordable Care Act (Employer Mandate): 1095-C")
        window_label.place(anchor = CENTER, relx = .5, rely = 0.081)
        window_label.config(font = ("Segoe UI", 10))

        find_file_label = Label(text = "Select a file to open")
        find_file_label.place(anchor = CENTER, relx = .5, rely = 0.12)
        browse_button = ttk.Button(self.window, text ="Browse", command = self.select_file)
        browse_button.place(relx = .46, rely = 0.145)


        #Tree to contain list of company employees
        self.tree = ttk.Treeview(self.window, height=23)
        self.tree['columns'] = ("Employee ID", "First Name", "Last Name")
        self.tree.column("#0", width=0, stretch=NO)
        self.tree.column("Employee ID", anchor=W, width=60)
        self.tree.column("First Name", anchor=W, width=100)
        self.tree.column("Last Name", anchor=W, width=100)
        self.tree.heading("#0", text="", anchor=W)
        self.tree.heading("Employee ID", text = "ID", anchor=W)
        self.tree.heading("First Name", text = "First Name", anchor=W)
        self.tree.heading("Last Name", text = "Last Name", anchor=W)
        self.tree.bind('<Motion>', 'break')
        self.tree.place(relx = .01, rely = 0.2)
        self.tree_index = 0
        self.tree.bind('<ButtonRelease-1>', self.selectItem)

        vsb = ttk.Scrollbar(self.window, orient="vertical", command=self.tree.yview)
        vsb.place(relx = .274, rely = 0.2, height=486)

        self.tree.configure(yscrollcommand=vsb.set)
        #TreeView Section Header
        
        listbox_label = Label(text = "Employee List",  font='Helvetica 9 bold')
        listbox_label.place(relx=0.01, rely = 0.162)

        #Canvas to hold both rectangles
        canvas = Canvas(width=670, height=510)
        canvas.place(relx = 0.32, rely = 0.187)
        canvas.create_rectangle(10,10,660,200, outline ="grey")
        canvas.create_rectangle(10,220,660,495, outline ="grey")  

        #General information section header
        general_label = Label(text = "General Information",  font='Helvetica 9 bold')
        general_label.place(relx = 0.34, rely = 0.187)

        #Widgets for company statistics
        self.n_employees = Label(text = "Number of Employees: ")
        self.n_employees.place(relx = 0.35, rely = 0.24)
        self.yhrs_average = Label(text = "Average hours worked [Yearly]: ")
        self.yhrs_average.place(relx = 0.35, rely = 0.28)
        self.mhrs_average = Label(text = "Average hours worked [Monthly]: ")
        self.mhrs_average.place(relx = 0.35, rely = 0.32)
        self.ft_months = Label(text = "Number of full-time months: ")
        self.ft_months.place(relx = 0.35, rely = 0.36)
        self.pt_months = Label(text = "Number of part-time months: ")
        self.pt_months.place(relx = 0.35, rely = 0.40)

        self.year_label = Label(text = "Filing Year: ")
        self.year_label.place(relx = 0.73, rely = 0.24)
        self.filing_state = Label(text = "Status of Filing: ")
        self.filing_state.place(relx = 0.73, rely = 0.28)
        self.filing_penalty = Label(text = "Penalty for failure to file: ")
        self.filing_penalty.place(relx = 0.73, rely = 0.32)


        #Employee information section header
        employee_info = Label(text = "Employee Information", font='Helvetica 9 bold')
        employee_info.place(relx = 0.34, rely = 0.509)

        #Widgets for search functionality
        search_header = Label(text = "Employee Search: ")
        search_header.place(relx = 0.35, rely = 0.544)
        self.promt_label = Label(text = "", fg = "red", font='Helvetica 8 bold')
        self.promt_label.place(relx = 0.54, rely = 0.544)
        id_search = Label(text = "ID: ")
        id_search.place(relx = 0.34, rely = 0.6)
        self.id_input = Entry(self.window)
        self.id_input.place(relx = 0.36, rely = 0.6)

        fname_search = Label(text = "First Name: ")
        fname_search.place(relx = 0.49, rely = 0.6)
        self.finput = Entry(self.window)
        self.finput.place(relx = 0.56, rely = 0.6)

        fname_search = Label(text = "Last Name: ")
        fname_search.place(relx = 0.69, rely = 0.6)
        self.linput = Entry(self.window)
        self.linput.place(relx = 0.76, rely = 0.6)

        search_button = ttk.Button(self.window, text ="Search", command = self.getSearch)
        search_button.place(relx = 0.90, rely = 0.594)

        #Widgets for employee statistics
        self.current_employee = Label(text = "ID: ")
        self.current_employee.place(relx = 0.35, rely = 0.67)    

        self.name = Label(text = "Name: ")
        self.name.place(relx = 0.35, rely = 0.72)  

        self.employee_hours = Label(text = "Hours worked: ")
        self.employee_hours.place(relx = 0.35, rely = 0.77) 

        self.employee_months = Label(text = "Months worked: ")
        self.employee_months.place(relx = 0.35, rely = 0.82) 

        self.jan = Label(text = "January: ")
        self.jan.place(relx = 0.56, rely = 0.67)
        
        self.feb = Label(text = "February: ")
        self.feb.place(relx = 0.56, rely = 0.72)

        self.march = Label(text = "March: ")
        self.march.place(relx = 0.56, rely = 0.77)

        self.april = Label(text = "April: ")
        self.april.place(relx = 0.56, rely = 0.82)

        self.may = Label(text = "May: ")
        self.may.place(relx = 0.70, rely = 0.67)

        self.june = Label(text = "June: ")
        self.june.place(relx = 0.70, rely = 0.72)

        self.july = Label(text = "July: ")
        self.july.place(relx = 0.70, rely = 0.77)

        self.august = Label(text = "August: ")
        self.august.place(relx = 0.70, rely = 0.82)

        self.sep = Label(text = "September: ")
        self.sep.place(relx = 0.84, rely = 0.67)

        self.oct = Label(text = "Ocotober: ")
        self.oct.place(relx = 0.84, rely = 0.72)

        self.nov = Label(text = "November: ")
        self.nov.place(relx = 0.84, rely = 0.77)

        self.dec = Label(text = "December: ")
        self.dec.place(relx = 0.84, rely = 0.82)

        #Helper lists to allow easy access to labels
        self.month_labels = [self.jan, self.feb, self.march, self.april, self.may, self.june, self.july, self.august, self.sep, self.oct, self.nov, self.dec]
        self.month_text = ["January: ", "February: ", "March: ", "April: ", "May: ", "June: ", "July: ", "August: ", "September: ", "October: ", "November: ", "December: ",]


    
    #Function for TreeView onclick event
    def selectItem(self, a) -> None:
        #Get number of items in TreeView
        item_count = len(self.tree.get_children())
        if item_count != 0:
            #Clear out any previous error labels
            self.promt_label.config(text = "")
            #Get selected item in treeview
            curItem = self.tree.focus()
            id = self.tree.item(curItem)["values"][0]
            #Pupulate employee statistics
            empl_set = self.set_empl_statistics(id)
            if empl_set == False:
                print("Employee Set Statistics Error")
                return
    
    #Function to populate employee statistics using ID
    def set_empl_statistics(self, id) -> bool:
        #Get list of employees
        employee_list = self.company.getEmployees()
        employee = employee_list.get(id)
        if employee:
            self.current_employee.config(text = "ID: " + str(id))
            self.name.config(text = "Name: " + employee.getFirstName() + " " + employee.getLastName())
            self.employee_hours.config(text = "Hours Worked: " + str(employee.getTotalHours()))
            self.employee_months.config(text = "Months Worked: " + str(employee.getTotalMonths()))
            months = employee.getMonthlyHours()
            for index, item in enumerate(months):
                if months[index] is None or months[index] == 0:
                    self.month_labels[index].config(text = self.month_text[index] + "N/A")
                elif months[index] < 130:
                    self.month_labels[index].config(text = self.month_text[index] + "PT")
                else:
                    self.month_labels[index].config(text = self.month_text[index] + "FT") 
            return True

        return False

    def select_file(self) -> None:
        #Clear out any previous error labels
        self.promt_label.config(text = "")
        #Open file dialog
        file_name = filedialog.askopenfilename(initialdir="/", title="Select ACA File", filetypes=(("xlsx files","*.xlsx"),))
        self.company = Company(file_name)
        company_employees = self.company.getEmployees()
        #Clear out tree
        self.tree.delete(*self.tree.get_children())
        #Add employees to TreeView
        for id, employee in company_employees.items():
            self.tree.insert(parent='', index ='end', iid=self.tree_index, values=(id, employee.getFirstName(), employee.getLastName()))
            self.tree_index += 1
        #Set company statistics
        comp_set = self.set_comp_statistics()
        if comp_set == False:
            print("Error setting company statistics")

    #Function to set company statistics
    def set_comp_statistics(self) -> bool:
        company_employees = self.company.getEmployees()
        #Error if no employees in company
        if len(company_employees) == 0:
            print("No Employees Found")
            return False
        #Setting the filing status labels
        filing_year = int("20" + str(self.company.getYear()))
        todays_date = date.today()
        if (filing_year == todays_date.year and (filing_year == todays_date.year and todays_date.month < 3)) or filing_year > todays_date.year:
            self.filing_state.config(text = "Status of Filing: OPEN", fg = "green")
            self.year_label.config(text = "Filing Year: " + str(filing_year), fg = "green")
            self.filing_penalty.config(text = "Penalty for failure to file: $" + str(len(company_employees) * 280), fg = "green")
        else:
            self.filing_state.config(text = "Status of Filing: LATE", fg = "red")
            self.year_label.config(text = "Filing Year: " + str(filing_year), fg = "red")
            self.filing_penalty.config(text = "Penalty for failure to file: $" + str(len(company_employees) * 280), fg = "red")

        #Getting & Setting employee related statistics
        num_employees, num_ftmonths, num_ptmonths, total_hours = len(company_employees), 0, 0, 0
        for id, employee in company_employees.items():
            months = employee.getMonthlyHours()
            for index, hrs in enumerate(months):
                if hrs is None or hrs == 0:
                    continue
                elif hrs < 130:
                    num_ptmonths += 1
                    total_hours += hrs
                else:
                    num_ftmonths += 1
                    total_hours += hrs

        self.n_employees.config(text = "Number of Employees: " + str(num_employees))
        self.yhrs_average.config(text = "Average hours worked [Yearly]: " + str(round(total_hours/num_employees)))
        self.mhrs_average.config(text = "Average hours worked [Monthly]: " + str(round((total_hours/num_employees)/12)))
        self.ft_months.config(text = "Number of full-time months: " + str(num_ftmonths))
        self.pt_months.config(text = "Number of part-time months: " + str(num_ptmonths))
        
        return True

    def getSearch(self) -> None:
        #Return if company is none
        if self.company is None:
            return
        #Reset search prompt
        self.promt_label.config(text = "")
        id = self.id_input.get()
        #Search will be case insensitive
        fname = self.finput.get().upper()
        lname = self.linput.get().upper()
        #Reset inputs
        self.id_input.delete(0, tk.END)
        self.finput.delete(0, tk.END)
        self.linput.delete(0, tk.END)
        found = False

        employees_dict = self.company.getEmployees()
        
        #If no employees in company
        if len(employees_dict) == 0:
            self.promt_label.config(text = "NO EMPLOYEES TO SEARCH")
            return

        #If given ID, find employee and compare to first and last name if given
        if id != "":
            #Return early if ID is not numeric
            if not id.isnumeric():
                self.promt_label.config(text = "ID MUST BE NUMERIC")
                return
            
            id = int(id)
            employee = employees_dict.get(id)
            if fname == "" and lname == "":
                found = self.set_empl_statistics(id)
            elif fname != "":
                if fname == employee.getFirstName():
                    found = self.set_empl_statistics(id)
            else:
                if lname == employee.getLastName():
                    found = self.set_empl_statistics(id)
        #If not given ID, find employee with the given first name and last name
        elif fname != "" and lname != "":
            for employee_id in employees_dict:
                if employees_dict[employee_id].getFirstName().upper() == fname and employees_dict[employee_id].getLastName().upper() == lname:
                    self.set_empl_statistics(employee_id)
                    found = True
                    break
        #Default Error -> 
        else:
            self.promt_label.config(text = "Please input a numeric ID or First and Last name")
            #Set found to True so following if condition does not overwrite prompt_label
            found = True
        #If the Employee is found
        if not found:
            self.promt_label.config(text = "EMPLOYEE NOT FOUND")

#Contains Company information for easy access
class Company:
    #Read excel file and save the necessary information
    def __init__(self, text_file) -> None:
        dataframe = openpyxl.load_workbook(text_file)
        self.dataframe1 = dataframe.active
        non_numeric_lst = ["01","02","03","04","05","06","07","08","09"]
        #Dictionary that will hold key:value pairs that consist of -> ID:Employee Object
        self.employee_dict = {}
        self.date_year = None
        #Regex used to check if a string is of the format DD/MM/YYYY or DD/MM/YY
        pattern = re.compile("^(1[0-2]|0?[1-9])/(3[01]|[12][0-9]|0?[1-9])/(?:[0-9]{2})?[0-9]{2}$")

        #Get the year of the date from the second column
        for col in self.dataframe1.iter_cols(2, 2):
            for row in range(0, self.dataframe1.max_row):
                date_str = col[row].value
                if isinstance(date_str, str):
                    if (pattern.match(date_str)):
                        self.date_year = date_str[date_str.rfind('/')+1:]
                        break
        #Find each employee by ID (first row), then find all the information associated with that ID and create employee object
        for col in self.dataframe1.iter_cols(1, 1):
            for row in range(0, self.dataframe1.max_row):
                #List that will hold the number of hours worked each month
                months_info = [None]*12
                employee_name = None
                id = None
                if isinstance(col[row].value, int) or col[row].value in non_numeric_lst:
                    total_hours, total_months = 0, 0
                    id = col[row].value
                    if id in non_numeric_lst:
                        id = int(id[1:])
                    #Get employee name from column 4
                    for col2 in self.dataframe1.iter_cols(4, 4):
                        employee_name = self.getCell(row, 4).split()
                    #Get total for each month if found
                    for row2 in range(row+1, self.dataframe1.max_row):
                        if col[row2].value == "YTD":
                            break
                        elif col[row2].value == "JANUARY":
                            months_info[0] = self.getCell(row2, 4)
                            if months_info[0] != 0:
                                total_hours += months_info[0]
                                total_months += 1
                        elif col[row2].value == "FEBUARY":
                            months_info[1] = self.getCell(row2, 4)
                            if months_info[1] != 0:
                                total_hours += months_info[1]
                                total_months += 1
                        elif col[row2].value == "MARCH":
                            months_info[2] = self.getCell(row2, 4)
                            if months_info[2] != 0:
                                total_hours += months_info[2]
                                total_months += 1
                        elif col[row2].value == "APRIL":
                            months_info[3] = self.getCell(row2, 4)
                            if months_info[3] != 0:
                                total_hours += months_info[3]
                                total_months += 1
                        elif col[row2].value == "MAY":
                            months_info[4] = self.getCell(row2, 4)
                            if months_info[4] != 0:
                                total_hours += months_info[4]
                                total_months += 1
                        elif col[row2].value == "JUNE":
                            months_info[5] = self.getCell(row2, 4)
                            if months_info[5] != 0:
                                total_hours += months_info[5]
                                total_months += 1
                        elif col[row2].value == "JULY":
                            months_info[6] = self.getCell(row2, 4)
                            if months_info[6] != 0:
                                total_hours += months_info[6]
                                total_months += 1
                        elif col[row2].value == "AUGUST":
                            months_info[7] = self.getCell(row2, 4)
                            if months_info[7] != 0:
                                total_hours += months_info[7]
                                total_months += 1
                        elif col[row2].value == "SETEMBER":
                            months_info[8] = self.getCell(row2, 4)
                            if months_info[8] != 0:
                                total_hours += months_info[8]
                                total_months += 1
                        elif col[row2].value == "OCTOBER":
                            months_info[9] = self.getCell(row2, 4)
                            if months_info[9] != 0:
                                total_hours += months_info[9]
                                total_months += 1
                        elif col[row2].value == "NOVEMBER":
                            months_info[10] = self.getCell(row2, 4)
                            if months_info[10] != 0:
                                total_hours += months_info[10]
                                total_months += 1
                        elif col[row2].value == "DECEMBER":
                            months_info[11] = self.getCell(row2, 4)
                            if months_info[11] != 0:
                                total_hours += months_info[11]
                                total_months += 1
                    #Create the employee object and add it to dictionary
                    if id is not None:
                        #Create tuple that contains the necessary information to create employee object
                        employee_info = (id, employee_name, months_info, total_months, total_hours)
                        employee = Employee(employee_info)
                        self.employee_dict[employee.getId()] = employee

    #Helper function to return value at given row and column
    def getCell(self, r, c) -> str:
        for row in range(r, r+1):
            for col2 in self.dataframe1.iter_cols(c, c):
                return col2[row].value

    #Returns dictionary containing all the employees in company
    def getEmployees(self) -> dict:
        return self.employee_dict
    
    #Returns the year of the data
    def getYear(self) -> str:
        return self.date_year


#Employee class to allow creation of employee objects which encapsulates each employee's information
class Employee:
    #initialize employee object
    def __init__(self, employee_info) -> None:
        self.id = employee_info[0]
        self.first_name = employee_info[1][1]
        self.last_name = employee_info[1][0][:-1]
        self.month_values = employee_info[2]
        self.total_months = employee_info[3]
        self.total_hours = employee_info[4]
    
    #Return employee id
    def getId(self) -> int:
        return self.id
    
    #Return first name
    def getFirstName(self) -> str:
        return self.first_name

    #Return last name
    def getLastName(self) -> str:
        return self.last_name
    
    #Return list containing hours worked each month, indexed 0-11
    def getMonthlyHours(self) -> list:
        return self.month_values

    #Return total hours worked
    def getTotalHours(self) -> int:
        return self.total_hours

    #Return total number of months worked
    def getTotalMonths(self) -> int:
        return self.total_months


def main(): 
    #Start tkinter window
    root = Tk()
    ma = MainApplication(root)
    root.mainloop()

if __name__ == '__main__':
    main()


        



